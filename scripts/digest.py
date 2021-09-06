#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Created on Sat Sep  4 12:34:05 2021

@author: nerohmot
"""
import sys
import os
import yaml
import requests
import json
import bz2
import io
import tarfile
import openpyxl
import argparse

supported_subdirs = ['linux-64', 'linux-aarch64', 'win-64', 'osx-64', 'osx-arm64']

def create_digest(args):

    def as_text(value):
        if value is None:
            return ""
        return str(value)

    def next_xlsx_column(col_str):
        last_char = col_str[-1]
        if last_char == "Z":
            if col_str[:-1].count("Z") == len(col_str[:-1]):
                return "A" * (len(col_str) + 1)
            else:
                return f"{col_str[:-2]}{chr(ord(col_str[-2])+1)}A"
        else:
            return f"{col_str[:-1]}{chr(ord(last_char)+1)}"
        
    build_version = args.build    
    repo_root = os.path.dirname(os.path.dirname(os.path.normpath(__file__)))
    specs_fpath = os.path.join(repo_root, "specs.yaml")
    wb_name = os.path.join(repo_root, 'maxiconda-envs.xlsx')

    print(f"Creating digest for V{build_version}")

    data = {}
    with open(specs_fpath) as fd:
        specs = yaml.load(fd, Loader=yaml.FullLoader)
        environments = list(specs['environments'])  # list of all environments
        pfms = list(specs['matrix'])  # list of all platforms
        primary_packages = {}  # dictionary with all primary packages per environmnet
        for environment in environments:
            primary_packages[environment] = specs['environments'][environment]
        targets = []  # list of all targets
        column = {}  # dictionary, key = target & value = (xlsx_column, OS, CPU, PY) 
        last_column = "A"
        for pfm in pfms:
            for target in specs['matrix'][pfm]:
                if target not in targets:
                    targets.append(target)
                OS_CPU, PY = target.split("_")
                OS, CPU = OS_CPU.split("-")
                if OS == "win":
                    OS = "Windows"
                    if CPU == "64":
                        CPU = "x86_64"
                elif OS == "linux":
                    OS = "Linux"
                    if CPU == "64":
                        CPU = "x86_64"
                elif OS == "osx":
                    OS = "MacOS"
                    if CPU == "64":
                        CPU = "x86_64"
                else:
                    raise Exception(f"OS '{OS}' not implemented")
                
                this_column = next_xlsx_column(last_column)
                column[target] = (this_column, OS, CPU, PY)
                last_column = this_column                
        max_column = next_xlsx_column(last_column)
        for environment in environments:
            data[environment] = {}
            for target in targets:
                platform, python = target.split("_")
                if environment in specs['matrix'][platform][target]:
                    url = f"https://anaconda.org/Semi-ATE/{environment}/{build_version}/download/{platform}/{environment}-{build_version}-{python}.tar.bz2"
                    print(f"Analyzing '{url}' ... ", end="", flush=True)
                    request = requests.get(url)
                    if request.status_code != 200:
                        print("Skiped (file doesn't exist)")
                        continue
                    try:
                        json_file_tar_flo_blo_bz2 = request.content
                        json_file_tar_flo_blo = bz2.decompress(json_file_tar_flo_blo_bz2)
                        json_file_tar_flo = io.BytesIO(json_file_tar_flo_blo)
                        json_file_tar = tarfile.open(fileobj=json_file_tar_flo)
                        json_file_tar.extract("info/index.json", path=".")
                        with open("./info/index.json", 'r') as fd:
                            index = json.load(fd)
                    except:
                        print("Skipped (encountered a problem)")
                        continue
                    print("Done.")
                    if platform not in data[environment]:
                        data[environment][platform] = {}
                    if python not in data[environment][platform]:
                        data[environment][platform][python] = index['depends']

    print("Consolidating data ... ", end="", flush=True)

    # dictionary with all packages per environment
    all_packages = {}
    for environment in data:
        all_packages[environment] = []
        for platform in data[environment]:
            for python in data[environment][platform]:
                for package in data[environment][platform][python]:
                    name, version, build = package.split(' ')
                    if name not in all_packages[environment]:
                        all_packages[environment].append(name)

    # dictionary per enviroment where key = package name & key = row number
    row = {}
    wb = openpyxl.Workbook()
    for environment in data:
        row[environment] = {}
        wb.create_sheet(environment)
        ws = wb[environment]
        ws_dim_holder = openpyxl.worksheet.dimensions.DimensionHolder(worksheet=ws)
        ws["A1"] = f"V{build_version}"
        ws["A1"].font = openpyxl.styles.Font(bold=True, color='00FF0000')
        ws["A2"] = f"specs.yaml"  
        ws['A2'].hyperlink = f"https://github.com/Semi-ATE/maxiconda-envs/releases/download/{build_version}/specs.yaml"
        ws['A2'].style = "Hyperlink"

        # Insert the packages
        row_offset = 4
        for index, package in enumerate(sorted(all_packages[environment])):
            this_row = row_offset + index
            this_cell = f"A{this_row}"
            ws[this_cell] = package
            row[environment][package] = this_row
            # set the background of the primary packages to red
            if package in primary_packages[environment] or package == "python":
                last_column = "A"
                while last_column != max_column:
                    ws[f"{last_column}{this_row}"].fill = openpyxl.styles.PatternFill(start_color="D3D3D3", fill_type = "solid")
                    last_column = next_xlsx_column(last_column)
                
        # Insert the data per target
        for target in targets:
            this_column, OS, CPU, PY = column[target]
            platform = target.split("_")[0]
            ws[f"{this_column}1"] = OS 
            ws[f"{this_column}1"].alignment = openpyxl.styles.Alignment(horizontal='center')
            ws[f"{this_column}2"] = CPU
            ws[f"{this_column}2"].alignment = openpyxl.styles.Alignment(horizontal='center')
            ws[f"{this_column}3"] = PY
            ws[f"{this_column}3"].alignment = openpyxl.styles.Alignment(horizontal='center')
            for package in all_packages[environment]:
                this_row = row[environment][package]
                package_version = None
                package_build = None
                if environment in data:
                    if platform in data[environment]:
                        if PY in data[environment][platform]:
                            for env_package in data[environment][platform][PY]:
                                if env_package.startswith(package):
                                    _, package_version, package_build = env_package.split(" ")
                                    package_version = package_version.replace("=", "").strip()
                                    package_build = package_build.replace("=", "").strip()
                                    ws[f"{this_column}{this_row}"] = package_version.replace("=", "").strip()
                                    ws[f"{this_column}{this_row}"].comment = openpyxl.comments.Comment(package_build, "Semi-ATE")
                                    ws[f"{this_column}{this_row}"].alignment = openpyxl.styles.Alignment(horizontal='center')

        # adjust column width to fit ALL the column contenses
        for column_cells in ws.columns:
            new_column_length = max(len(as_text(cell.value)) for cell in column_cells)
            new_column_letter = (openpyxl.utils.get_column_letter(column_cells[0].column))
            if new_column_length > 0:
                ws.column_dimensions[new_column_letter].width = new_column_length + 1

    del wb['Sheet']
    print("Done.")
    print(f"Saving '{wb_name}' ... ", end="", flush=True)
    try:
        wb.save(wb_name)
    except:
        print("Failed!")
        sys.exit(1)
    else:
        print("Done.")
        sys.exit(0)

if __name__ == '__main__':
    parser = argparse.ArgumentParser()
    parser.add_argument('--build', default=os.getenv('MAXICONDA_ENV_RELEASE', "0.0.0"))
    args = parser.parse_args()
    
    create_digest(args)
    
