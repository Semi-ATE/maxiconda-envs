#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
This script can :
  - solve an environment (from any os to any other os)
  - build an environment (needs to run on the actual environment)
  - upload the build environment to the anaconda.org/Semi-ATE channel
"""
import argparse
import sys
import os
import platform
import subprocess
import shutil
import tempfile
import jinja2
import time
import datetime
import yaml
import requests
import bz2
import json
import tarfile
import io
import openpyxl
from openpyxl.worksheet.dimensions import ColumnDimension, DimensionHolder
from openpyxl.utils import get_column_letter
from pathlib import Path
from typing import Tuple

# Constants
HERE = Path(__file__).resolve().parent
REPO_ROOT = HERE.parent
SPECS_FPATH = REPO_ROOT / "specs.yaml"
RECIPES_ROOT  = REPO_ROOT / "recipes"
CACHE = {}

def get_subdir():
    """
    This function will return the OS and CPU in a conda-forge compatible format (aka: 'CONDA_SUBDIR')

    Parameters
    ----------
    /

    Returns
    -------
    conda-forge compatile OS_CPU, limited to:
        - linux-64   
        - linux-aarch64
        - osx-64   
        - osx-arm64
        - win-64
    """

    is_64bits = sys.maxsize > 2**32
    if not is_64bits:
        raise Exception("only 64 bit platforms are supported.")

    CPU = platform.machine()

    OS = platform.system()
    if OS == "Linux":
        if CPU in ["AMD64", "x86_64"]:
            OS_CPU = "linux-64"
        elif CPU in ["aarch64"]:
            OS_CPU = "linux-aarch64"
        else:
            raise Exception(f"'{CPU}' not supported in Linux")
    elif OS == "Windows":
        if CPU in ["AMD64", "x86_64"]:
            OS_CPU = "win-64"
        else:
            raise Exception(f"'{CPU}' not supported in Windows")
    elif OS == "Darwin":
        if CPU in ["AMD64", "x86_64"]:
            OS_CPU = "osx-64"
        elif CPU in ["aarch64"]:
            OS_CPU = "osx-arm64"
        else:
            raise Exception(f"'{CPU}' not supported in MacOS")
    else:
        raise Exception("'{OS}' not supported.")

    return OS_CPU

def upload(package_path):
    """
    This function uploads the package pointed to by package_path to anaconda.org/Semi-ATE
    It uses the CONDA_UPLOAD_TOKEN environment variable to do so.

    Parameters
    ----------
    package_path : str that points to a valid meta.yaml file under recipes.

    Returns
    -------
    bool (True for success, False for failure)
    """

    if not os.path.exists(package_path):
        raise Exception(f"'{package_path}' does not exist!")

    package_path = os.path.normpath(package_path)
    package_file = package_path.split(os.sep)[-1]
    environment = package_file.split("-")[0]
    package_version = package_file.split("-")[1]
    PY = package_file.split("-")[2].split(".")[0]
    OS_CPU = package_path.split(os.sep)[-2] 

    print(f"Uploading : {OS_CPU}/{PY}/{environment}")
    retval = True
    cmd = ["anaconda", "-t", os.environ.get("CONDA_UPLOAD_TOKEN", "Woops"), "upload", "-u", "semi-ate", package_path, "--force"]
    print(f"  running '{' '.join(cmd)}' ... ", end="", flush=True)
    p = subprocess.Popen(cmd, stdout=subprocess.PIPE, stderr=subprocess.PIPE)
    _, output = p.communicate()
    output_lines = output.decode("utf-8").split("\n")  
    for output_line in output_lines:
        if "[ERROR]" in output_line:
            retval = False
    if retval == False:
        print("Failure.")
        for output_line in output_lines:
            print(f"  {output_line}")
    else:
        print("Success.")
    return retval

def build(env_meta_path):
    """
    Builds the package for `env_meta_path`.


    Parameters
    ----------
    env_meta_path : str
        path pointing to a 'meta.yaml' file.

    Returns
    -------
    success : the path to the build package
    failure : None
    """

    if not os.path.exists(env_meta_path):
        raise Exception(f"'{env_meta_path}' does not exist!")

    env_meta_root = os.path.dirname(env_meta_path)
    if not os.path.exists(env_meta_root):
        raise Exception(f"'{env_meta_root}' does not exist!")

    env_meta_path = os.path.normpath(env_meta_path)
    environment = env_meta_path.split(os.sep)[-2]
    PY = env_meta_path.split(os.sep)[-3]
    OS_CPU = env_meta_path.split(os.sep)[-4]

    NUMPY_VER = None
    PYTHON_VER = None
    with open(env_meta_path) as fd:
        for line in fd:
            if "- numpy " in line:
                NUMPY_VER = line.split("=")[1].strip()
            if "- python " in line and PYTHON_VER is None:
                PYTHON_VER = line.split("=")[1].strip()

    cmd = ["mamba", "build", ".", "--python",  PYTHON_VER]
    # cmd = ["conda-build", ".", "--python",  PYTHON_VER]
    if NUMPY_VER:
        cmd.extend(["--numpy", NUMPY_VER])
    # cmd.extend([f"-c", "conda-forge", "-c", "Semi-ATE"])
    cmd.extend([f"-c", "conda-forge"])

    print(f"  running '{' '.join(cmd)}' in '{env_meta_root}' ... ", end="", flush=True)
    p = subprocess.Popen(cmd, cwd=env_meta_root, stdout=subprocess.PIPE, stderr=subprocess.PIPE)
    output, error = p.communicate()

    retval = None
    if b"anaconda upload" in output:
        print("success.")
        retval = output.split(b'anaconda upload \\')[1].replace(b"\r",b"").split(b'\n')[1].decode("utf-8").strip()
        print(f"  package location = '{retval}'")
    else:
        print("Failure.")
        output_lines = output.decode("utf-8").split("\n")
        error_lines = error.decode("utf-8").split("\n")
        for line in output_lines:
            print(line)
        for line in error_lines:
            print(line)
    return retval


def xbuild(env_meta_path, output_path=None):
    """
    Builds the package for `env_meta_path`.

    Parameters
    ----------
    env_meta_path : str
        path pointing to a 'meta.yaml' file.

    Returns
    -------
    success : the path to the build package
    failure : None
    """

    if not os.path.exists(env_meta_path):
        raise Exception(f"'{env_meta_path}' does not exist!")

    env_meta_root = os.path.dirname(env_meta_path)
    if not os.path.exists(env_meta_root):
        raise Exception(f"'{env_meta_root}' does not exist!")
    env_meta_path = os.path.normpath(env_meta_path)

    licens_path = os.path.join(env_meta_path)

    VERSION = os.getenv('MAXICONDA_ENV_RELEASE', "0.0.0")

    environment = env_meta_path.split(os.sep)[-2]
    PY = env_meta_path.split(os.sep)[-3]
    OS_CPU = env_meta_path.split(os.sep)[-4]
    OS, CPU = OS_CPU.split('-')
    if CPU == "64":
        CPU = "x86_64"
    NUMPY_VER = "1.21.2"
    PYTHON_VER = None
    DEPENDS = []
    deps = False
    MAINTAINERS = []
    maintainers = False
    with open(env_meta_path) as fd:
        for line in fd:
            line = line.lstrip()
            if line.startswith("number:"):
                BUILD_NUMBER = line.split(":")[-1].strip()
                deps = False
                maintainers = False
            if line.startswith("string:"):
                BUILD_STRING = line.split(":")[-1].strip()
                deps = False
                maintainers = False
            if line.startswith("home:"):
                HOME_URL = line.split(":")[1].strip()
                deps = False
                maintainers = False
            if line.startswith("license:"):
                LICENSE = line.split(":")[1].strip()
                deps = False
                maintainers = False
            if line.startswith("license_file:"):
                LICENSE_FILE = line.split(":")[1].strip()
                deps = False
                maintainers = False
            if line.startswith("summary:"):
                SUMMARY = line.split(":")[1].strip()
                deps = False
                maintainers = False
            if line.startswith("dev_url:"):
                DEV_URL = line.split(":")[1].strip()
                deps = False
                maintainers = False
            if line.startswith("about:"):
                deps = False
                maintainers = False
            if line.startswith("run:"):
                deps = True
                maintainers = False
            if line.startswith("maintainers:"):
                deps = False
                maintainers = True
            if deps: 
                if line.startswith("- "):
                    DEPENDS.append(line[1:].strip())
                if line.startswith("- numpy"):
                    NUMPY_VER = line.split("=")[1].strip()
                if line.startswith("- python"):
                    PYTHON_VER = line.split("=")[1].strip()
            if maintainers:
                if line.startswitn("- "):
                    MAINTAINERS.append(line[1:].strip())

    if output_path is None:
        archive_name = os.path.join(os.getcwd(), f"{environment}-{VERSION}-{PY}.tar.bz2")
    else:
        archive_name = os.path.join(output_path, f"{environment}-{VERSION}-{PY}.tar.bz2") 

    license_path = os.path.normpath(os.path.join(os.path.dirname(env_meta_path), LICENSE_FILE))

    index = {
        "arch" : CPU,
        "build" : BUILD_STRING,
        "build_number" : BUILD_NUMBER,
        "depends" : DEPENDS,
        "license": "MIT",
        "name": environment,
        "platform": OS,
        "subdir": OS_CPU,
        "timestamp": int(time.time()*1000),
        "version": VERSION
    }

    about = {
        "channels": [
            "https://conda.anaconda.org/Semi-ATE",
            "https://conda.anaconda.org/conda-forge"
         ],
        "conda_build_version": "3.21.4",
        "conda_private": False,
        "conda_version": "4.10.3",
        "dev_url": DEV_URL,
        "env_vars": {
            "CIO_TEST": "<not set>"
        },
        "extra": {
            "copy_test_source_files": True,
            "final": True,
            "recipe-maintainers": MAINTAINERS,
        },
        "home": HOME_URL,
        "identifiers": [],
        "keywords": [],
        "license": LICENSE,
        "license_file": LICENSE_FILE,
        "root_pkgs": DEPENDS,
        "summary": SUMMARY,
        "tags": []
    }

    PATHS = {
        "paths": [],
        "paths_version": 1
    }

    CONDA_BUILD_CONFIG = f"""c_compiler: gcc
cpu_optimization_target: nocona
cran_mirror: https://cran.r-project.org
cxx_compiler: gxx
extend_keys:
- ignore_version
- pin_run_as_build
- ignore_build_only_deps
- extend_keys
fortran_compiler: gfortran
ignore_build_only_deps:
- numpy
- python
lua: '5'
numpy: {NUMPY_VER}
perl: 5.26.2
pin_run_as_build:
  python:
    min_pin: x.x
    max_pin: x.x
  r-base:
    min_pin: x.x
    max_pin: x.x
python: {PYTHON_VER}
r_base: '3.5'
target_platform: {OS_CPU}
"""

    with tempfile.TemporaryDirectory() as tmpdirname:
        licenses_directory = os.path.join(tmpdirname, "licenses")
        os.makedirs(licenses_directory, mode=0o777, exist_ok=True)
        shutil.copyfile(license_path, os.path.join(tmpdirname, "licenses", "LICENSE0.txt"))

        recipe_directory = os.path.join(tmpdirname, "recipe")
        os.makedirs(recipe_directory, mode=0o777, exist_ok=True)
        shutil.copyfile(env_meta_path, os.path.join(tmpdirname, "recipe", "meta.yaml.template"))
        with open(os.path.join(tmpdirname, "recipe", "conda_build_config.yaml"), "w", encoding="utf-8") as file:
            file.write(CONDA_BUILD_CONFIG)
        file_loader = jinja2.FileSystemLoader(os.path.dirname(env_meta_path))
        env = jinja2.Environment(loader=file_loader)
        template = env.get_template('meta.yaml')
        env.globals.update(os=os)
        output = '\n'.join(template.render().split('\n')[9:])
        LOCATION = os.path.sep.join(os.path.dirname(env_meta_path).split(os.path.sep)[-5:])
        DATE = datetime.datetime.now().strftime("%A %B %d %Y @ %H:%M:%S")
        with open(os.path.join(tmpdirname, "recipe", "meta.yaml"), "w") as fd:
            fd.write("# This file is created by maxiconda-env/xbuild\n")
            fd.write(f"# meta.yaml template originally from 'https://github.com/Semi-ATE/{LOCATION}', last modified {DATE}\n")
            fd.write("# ------------------------------------------------\n\n")
            fd.write(output)

        with open(os.path.join(tmpdirname, "about.json"), "w", encoding="utf-8") as file:
            json.dump(about, file, indent=2)
        with open(os.path.join(tmpdirname, "files"), "w"):
            pass
        with open(os.path.join(tmpdirname, "git"), "w"):
            pass
        with open(os.path.join(tmpdirname, "hash_input.json"), "w", encoding="utf-8") as file:
            json.dump({}, file, indent=2)
        with open(os.path.join(tmpdirname, "index.json"), "w", encoding="utf-8") as file:
            json.dump(index, file, indent=2)
        with open(os.path.join(tmpdirname, "paths.json"), "w", encoding="utf-8") as file:
            json.dump(PATHS, file, indent=2)

        with tarfile.open(archive_name, mode='w:bz2') as archive:
            archive.add(tmpdirname, recursive=True, arcname="info")

    return archive_name

def is_uploadable(package_path):
    """
    Determines if `package_path` is uploadable.

    if the version is '0.0.0' this means it is a test build, and that is not uploaded.
    if `package_path` doesn't point to an existing file, obviously it is not uploadable.
    `package_path needs to be a string.
    `package_path` must end with '.tar.bz2'.
    
    Parameters
    ----------
    package_path : str
        path to the package.

    Returns
    -------
    True : if all conditons are fulfilled..
    False : if one or more conditions are not fulfilled. 
    """
    if not isinstance(package_path, str):
        return False
    if not package_path.endswith(".tar.bz2"):
        return False
    package_path = os.path.normpath(package_path)
    if not os.path.exists(package_path):
        return False
    retval = os.path.basename(package_path)
    retval = retval.split("-")[1]
    retval = not (retval == "0.0.0")
    return retval

def get_repodata(url):
    """
    Get repodata from url poitning to anaconda.org bz2 files.

    If url has been already downloaded in a session, it will use the data from
    cache.

    Parameters
    ----------
    url : str
        URL to repodata.

    Returns
    -------
    dict
        Repodata dictionary.
    """
    global CACHE
    if url not in CACHE:
        request = requests.get(url)
        if request.status_code != 200:
            print(f"Warning: '{url}' doesn't exist!")
            return {}
        arch_json_bz2 = request.content
        arch_json = bz2.decompress(arch_json_bz2) 
        arch = json.loads(arch_json)
        CACHE[url] = arch
    else:
        arch = CACHE[url]

    return arch

def get_packages_from_channel(channel, platform):
    """
    This function returns all packages that exist for the given designator in a given channel

    Parameters
    ----------
    platform
      str formatted like so :
        linux-64 --> to get all python versions

    Returns
    -------
    dict
        packages and versions formatted like { package : [versions] }
    """
    with open(SPECS_FPATH) as fd:
        specs = yaml.load(fd, Loader=yaml.FullLoader)
        platforms = list(specs['matrix'])
    if platform not in platforms:
        return {}

    repodata_url = f"https://conda.anaconda.org/{channel}/{platform}/repodata.json.bz2"
    packages = get_repodata(repodata_url)['packages']
    return packages

def get_conda_forge_packages(designator):
    """
    This function returns all packages that exist for the given designator

    Parameters
    ----------
    designator 
        str formatted like so :linux-64_py36

    Returns
    -------
    dict
        packages and versions formatted like { package : [versions] }
    """
    OS_CPU = designator.split('_')[0]
    PY = designator.split('_')[1]

    arch_packages = f"https://conda.anaconda.org/conda-forge/{OS_CPU}/repodata.json.bz2"
    noarch_packages = "https://conda.anaconda.org/conda-forge/noarch/repodata.json.bz2"
    retval = {}

    arch = get_repodata(arch_packages)
    for package in arch['packages']:
        if arch['packages'][package]['build'].startswith('py'):
            if not arch['packages'][package]['build'].startswith(PY[-4:]):
                continue

        if arch['packages'][package]['name'] not in retval:
            retval[arch['packages'][package]['name']] = []

        if arch['packages'][package]['version'] not in retval[arch['packages'][package]['name']]:
            retval[arch['packages'][package]['name']].append(arch['packages'][package]['version'])

    noarch = get_repodata(noarch_packages)
    for package in noarch['packages']:
        if noarch['packages'][package]['name'] not in retval:
            retval[noarch['packages'][package]['name']] = []

        if noarch['packages'][package]['version'] not in retval[noarch['packages'][package]['name']]:
            retval[noarch['packages'][package]['name']].append(noarch['packages'][package]['version'])

    return retval

def reduce(packages, designator):
    """ 
    This function will return a list of packages that need to be removed from packages because :
        1) They don't exist on conda-forge for the given designator.
        2) the package is Python
        3) the package is pypy

    Parameters
    ----------
    packages : list of packages
    designator 
        str formatted like so :linux-64_py36

    Returns
    -------
    list of packages that do not exist for the given designator.
    """
    retval = []
    available_packages = get_conda_forge_packages(designator)
    for package in packages:
        if not package in available_packages:
            retval.append(package)
        if package.startswith("python"):
            retval.append(package)
        if package.startswith("pypy"):
            retval.append(package)

    return retval

def run_solver(pkgs, PY, channels=["conda-forge"], solver="mamba"):
    """
    Run conda solver dry run. 

    Parameters
    ----------
    pkgs : list of packages to solve for
    PY : string indicating the python version/implementation to use for solving
        py36 --> python=3.6
        pypy36 --> pypy3.6
    channels : list of channels to use
    solver : the solver to use

    Returns
    -------
    PY_IMP : string representing the python implementation
    data : The solution in (json) dictionary format.
    feedback : this is None on success, if not this is the "message"
    """
    if solver not in ["conda", "mamba"]:
        raise ValueError("Must use 'conda' or 'mamba' as solver!")

    if PY.startswith("pypy"):
        if not len(PY) == 6:
            raise ValueError(f"'{PY}' should be exactly 6 characters long")
        PY_IMP = f"pypy{PY[4]}.{PY[5]}"
    elif PY.startswith("py"):
        if not len(PY) == 4:
            raise ValueError(f"'{PY}' should be exactly 4 characters long")
        PY_IMP = f"python={PY[2]}.{PY[3]}"
    else:
        raise ValueError(f"'{PY}' should start with 'py' or 'pypy'")

    cmd = [solver, "create", "--name", "test_env", "--dry-run", "--json", "--yes", "--strict-channel-priority", PY_IMP] + pkgs
    for channel in channels:
        cmd.append("--channel")
        cmd.append(channel)

    print(f"  CONDA_SUBDIR = {os.environ.get('CONDA_SUBDIR', get_subdir())}")
    print("  solving command: '" + " ".join(cmd) + "'")

    p = subprocess.Popen(cmd, stdout=subprocess.PIPE, stderr=subprocess.PIPE)
    stdout, _ = p.communicate()
    data = {}
    if stdout.decode("utf-8").startswith("Encountered problems while solving."):
        feedback = stdout.decode("utf-8").split("Problem:")[1].strip()
    else:
        try:
            data = json.loads(stdout)
            feedback = None
        except Exception as err:
            feedback = f"Error: {err}"
            feedback += stdout.decode("utf-8")

    return PY_IMP, data, feedback

def solve(implementation, environment, with_buildstring=True):
    """
    Run conda solver dry run. 

    Parameters
    ----------
    implementation :
        str formatted like so :linux-64_py36
    environment :
        the name of the environment to write the recipe for

    Returns
    -------
    directory where the 'meta.yaml' is written, None if there was a problem.
    """
    OS_CPU, PY = implementation.split('_')

    with open(SPECS_FPATH) as fd:
        specs = yaml.load(fd, Loader=yaml.FullLoader)
        spec_primary_packages = specs['environments'][environment]

    print(f"  primary packages in spec : {spec_primary_packages}")

    packages_to_remove = reduce(spec_primary_packages, implementation)
    if packages_to_remove:
        print(f"  excluding packages : {packages_to_remove}")
    cleaned_primary_packages = []
    for package in spec_primary_packages:
        if package not in packages_to_remove:
            cleaned_primary_packages.append(package)
    if packages_to_remove:
        print(f"  remaining primary packages : {cleaned_primary_packages}")

    if not cleaned_primary_packages:
        print(f"  aborting : no packages to solve for.")
        return None

    PY_IMP, data, feedback = run_solver(cleaned_primary_packages, PY)

    if feedback:
        print(f"  Problem : {feedback}")
        return None

    all_packages = {}
    for element in data['actions']['LINK']:
        all_packages[element['name']] = {"version": element['version'], "build_string": element['build_string']}

    PYPY = None
    primary_packages = {}
    secondary_packages = {}
    for package in all_packages:
        if package == "python":
            PYTHON = f"{package} ={all_packages[package]['version']}"
            if with_buildstring:
                PYTHON += f" ={all_packages[package]['build_string']}"
        elif package.startswith("pypy"):
            PYPY = f"{package} ={all_packages[package]['version']}"
            if with_buildstring:
                PYPY += f" ={all_packages[package]['build_string']}"
        elif package in cleaned_primary_packages:
            primary_packages[package] = all_packages[package]
        else:
            secondary_packages[package] = all_packages[package]
    print(f"  primary packages : '{len(primary_packages)}'")
    print(f"  secondary packages : '{len(secondary_packages)}'")

    recipe_fpath = RECIPES_ROOT / OS_CPU / PY / environment / "meta.yaml"
    if not recipe_fpath.is_file():
        os.makedirs(recipe_fpath.parent, exist_ok=True)
    print(f"  writing : '{recipe_fpath}'")

    with open(recipe_fpath, 'w') as fh:
        fh.write("#\n# Copyright (c) Semi-ATE\n")
        fh.write("# Distributed under the terms of the MIT License\n")
        fh.write("#\n")
        fh.write(f"# {OS_CPU}/{PY}/{environment} (created on {get_subdir()})\n")
        fh.write("#\n\n")
        fh.write('{% set version = os.environ.get("MAXICONDA_ENV_RELEASE", "0.0.0") %}\n')
        fh.write("\n")
        fh.write("package:\n")
        fh.write(f"  name: {environment}\n")
        fh.write("  version: {{ version }}\n")
        fh.write("\n")
        fh.write("source:\n") 
        fh.write(f"  path: .\n")
        fh.write("\n")
        fh.write("build:\n")
        fh.write("  number: 0\n")
        fh.write(f"  string: {PY}\n")
        fh.write("\n")
        fh.write("requirements:\n")
        fh.write("  build:\n")
        fh.write(f"    - {PYTHON}\n")
        fh.write("  run:\n")
        fh.write(f"    - {PYTHON}\n")
        if PYPY:
            fh.write(f"    - {PYPY}\n")
        fh.write("\n")
        fh.write(f"    # {len(primary_packages)} primary packages :\n")
        for primary_package in sorted(primary_packages):
            fh.write(f"    - {primary_package} ={primary_packages[primary_package]['version']}")
            if with_buildstring:
                fh.write(f" ={primary_packages[primary_package]['build_string']}")
            fh.write("\n")
        fh.write("\n")
        fh.write(f"    # {len(secondary_packages)} secondary packages :\n")
        for secondary_package in sorted(secondary_packages):
            fh.write(f"    - {secondary_package} ={secondary_packages[secondary_package]['version']}")
            if with_buildstring:
                fh.write(f" ={secondary_packages[secondary_package]['build_string']}")
            fh.write("\n")
        fh.write("about:\n")
        fh.write("  home: https://github.com/Semi-ATE/maxiconda-envs\n")
        fh.write("  license: MIT\n")
        fh.write("  license_file: ../../../../LICENSE\n")
        fh.write(f"  summary: '{environment} meta package'\n")
        fh.write("  dev_url: https://github.com/Semi-ATE/maxiconda-envs\n\n")
        fh.write("extra:\n")
        fh.write("  recipe-maintainers:\n")
        fh.write("    - nerohmot\n")

    return str(recipe_fpath)

def next_xlsx_column(col_str):
    last_char = col_str[-1]
    if last_char == "Z":
        if col_str[:-1].count("Z") == len(col_str[:-1]):
            return "A" * (len(col_str) + 1)
        else:
            return f"{col_str[:-2]}{chr(ord(col_str[-2])+1)}A"
    else:
        return f"{col_str[:-1]}{chr(ord(last_char)+1)}"

def create_digest():

    def as_text(value):
        if value is None:
            return ""
        return str(value)

    release = os.environ.get("MAXICONDA_ENV_RELEASE", '0.0.27')
    print(f"Creating digest for V{release}")

    data = {}
    with open(SPECS_FPATH) as fd:
        specs = yaml.load(fd, Loader=yaml.FullLoader)
        environments = list(specs['environments'])  # list of all environments
        platforms = list(specs['matrix'])  # list of all platforms
        primary_packages = {}  # dictionary with all primary packages per environmnet
        for environment in environments:
            primary_packages[environment] = specs['environments'][environment]
        targets = []  # list of all targets
        column = {}  # dictionary, key = target & value = (xlsx_column, OS, CPU, PY) 
        last_column = "A"
        for platform in platforms:
            for target in specs['matrix'][platform]:
                if target not in targets:
                    targets.append(target)
                OS_CPU, PY = target.split("_")
                OS, CPU = OS_CPU.split("-")
                if OS == "win":
                    OS = "Windows"
                    if CPU == "64":
                        CPU = "x86_64"
                elif OS == "linux":
                    OS = "Linux"
                    if CPU == "64":
                        CPU = "x86_64"
                elif OS == "osx":
                    OS = "MacOS"
                    if CPU == "64":
                        CPU = "x86_64"
                else:
                    raise Exception(f"OS '{OS}' not implemented")
                
                this_column = next_xlsx_column(last_column)
                column[target] = (this_column, OS, CPU, PY)
                last_column = this_column                
        max_column = next_xlsx_column(last_column)
        for environment in environments:
            data[environment] = {}
            for target in targets:
                platform, python = target.split("_")
                if environment in specs['matrix'][platform][target]:
                    url = f"https://anaconda.org/Semi-ATE/{environment}/{release}/download/{platform}/{environment}-{release}-{python}.tar.bz2"
                    print(f"Analyzing '{url}' ... ", end="")
                    request = requests.get(url)
                    if request.status_code != 200:
                        print("Skiped (file doesn't exist)")
                        continue
                    try:
                        json_file_tar_flo_blo_bz2 = request.content
                        json_file_tar_flo_blo = bz2.decompress(json_file_tar_flo_blo_bz2)
                        json_file_tar_flo = io.BytesIO(json_file_tar_flo_blo)
                        json_file_tar = tarfile.open(fileobj=json_file_tar_flo)
                        json_file_tar.extract("info/index.json", path=f".")
                        with open("./info/index.json", 'r') as fd:
                            index = json.load(fd)
                    except:
                        print("Skipped (encountered a problem)")
                        continue
                    print("Done.")
                    if platform not in data[environment]:
                        data[environment][platform] = {}
                    if python not in data[environment][platform]:
                        data[environment][platform][python] = index['depends']

    print("Consolidating data ... ", end="", flush=True)

    # dictionary with all packages per environment
    all_packages = {}
    for environment in data:
        all_packages[environment] = []
        for platform in data[environment]:
            for python in data[environment][platform]:
                for package in data[environment][platform][python]:
                    name, version, build = package.split(' ')
                    if name not in all_packages[environment]:
                        all_packages[environment].append(name)

    # dictionary per enviroment where key = package name & key = row number
    row = {}
    wb = openpyxl.Workbook()
    wb_name = f"{str(REPO_ROOT / 'maxiconda-envs.xlsx')}"
    for environment in data:
        row[environment] = {}
        wb.create_sheet(environment)
        ws = wb[environment]
        ws_dim_holder = DimensionHolder(worksheet=ws)
        ws["A1"] = f"V{release}"
        ws["A1"].font = openpyxl.styles.Font(bold=True, color='00FF0000')

        # Insert the packages
        row_offset = 4
        for index, package in enumerate(sorted(all_packages[environment])):
            this_row = row_offset + index
            this_cell = f"A{this_row}"
            ws[this_cell] = package
            row[environment][package] = this_row
            # set the background of the primary packages to red
            if package in primary_packages[environment] or package == "python":
                last_column = "A"
                while last_column != max_column:
                    ws[f"{last_column}{this_row}"].fill = openpyxl.styles.PatternFill(start_color="D3D3D3", fill_type = "solid")
                    last_column = next_xlsx_column(last_column)


                
        # Insert the data per target
        for target in targets:
            this_column, OS, CPU, PY = column[target]
            platform = target.split("_")[0]
            ws[f"{this_column}1"] = OS 
            ws[f"{this_column}1"].alignment = openpyxl.styles.Alignment(horizontal='center')
            ws[f"{this_column}2"] = CPU
            ws[f"{this_column}2"].alignment = openpyxl.styles.Alignment(horizontal='center')
            ws[f"{this_column}3"] = PY
            ws[f"{this_column}3"].alignment = openpyxl.styles.Alignment(horizontal='center')
            for package in all_packages[environment]:
                this_row = row[environment][package]
                package_version = None
                package_build = None
                if environment in data:
                    if platform in data[environment]:
                        if PY in data[environment][platform]:
                            for env_package in data[environment][platform][PY]:
                                if env_package.startswith(package):
                                    _, package_version, package_build = env_package.split(" ")
                                if package_version and package_build:
                                    ws[f"{this_column}{this_row}"] = package_version
                                    ws[f"{this_column}{this_row}"].comment = openpyxl.comments.Comment(package_build, "Semi-ATE")
                                    ws[f"{this_column}{this_row}"].alignment = openpyxl.styles.Alignment(horizontal='center')

        # adjust column width to fit ALL the column contenses
        for column_cells in ws.columns:
            new_column_length = max(len(as_text(cell.value)) for cell in column_cells)
            new_column_letter = (openpyxl.utils.get_column_letter(column_cells[0].column))
            if new_column_length > 0:
                ws.column_dimensions[new_column_letter].width = new_column_length + 1

    del wb['Sheet']
    print("Done.")
    print(f"Saving '{wb_name}' ... ", end="", flush=True)
    try:
        wb.save(wb_name)
    except:
        print("Failed!")
    else:
        print("Done.")






    # order the info in an xlsx
    # save the xlsx under the name maxiconda-envs.xlsx

def main(args):

    if args.solve == args.build == args.digest == False:
        raise Exception("at least one action needs to be given (--solve --build --digest)")

    if args.upload and not args.build:
        print("uploading implies also building! (add --build)")
        args.build = True

    if args.digest:
        create_digest()
    else:
        CONDA_SUBDIR = os.environ.get("TARGET_CONDA_SUBDIR")
        if CONDA_SUBDIR is None:
            CONDA_SUBDIR = get_subdir()
            print(">>>>>>>>>>>>>>>> why I don't see TARGET_CONDA_SUBDIR environment variable ?!?")

        if os.path.exists(SPECS_FPATH):
            with open(SPECS_FPATH) as fd:
                specs = yaml.load(fd, Loader=yaml.FullLoader)
                implementations = specs['matrix'][CONDA_SUBDIR]
                environments = specs['environments']
        else:
            raise Exception(f"'{str({SPECS_FPATH})}' does not exits!")

        for implementation in implementations:
            for environment in implementations[implementation]:
                if environment not in environments:
                    raise Exception(f"Implementation '{implementation}' references the '{environment}' environment which is not defined.")
                if args.solve:
                    print(f"Solving : '{CONDA_SUBDIR}/{implementation.split('_')[1]}/{environment}' (from {get_subdir()})")
                    recipe_dir = solve(implementation, environment)
                else:
                    recipe_dir = str(RECIPES_ROOT / f"{CONDA_SUBDIR}/{implementation.split('_')[1]}/{environment}/meta.yaml")
                if args.build:
                    print(f"Building : '{CONDA_SUBDIR}/{implementation.split('_')[1]}/{environment}' ... ", end="")
                    package_path = xbuild(recipe_dir)
                    print(package_path)
                if args.upload:
                    if package_path:
                        if is_uploadable(package_path):
                            upload(package_path)
                        else:
                            print(f"Upload : Skipped (test build)")
                    else:
                        print(f"Upload : Skipped (build not successfull)")

if __name__ == '__main__':
    parser = argparse.ArgumentParser()
    parser.add_argument('--solve', action='store_true')
    parser.add_argument('--build', action='store_true')
    parser.add_argument('--upload', action='store_true')
    parser.add_argument('--digest', action='store_true')
    args = parser.parse_args()
    main(args)

